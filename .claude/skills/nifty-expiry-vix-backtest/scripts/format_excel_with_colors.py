#!/usr/bin/env python3
"""
Format VIX Backtest Results with Conditional Coloring

Applies green background to rows where VIX prediction was close but slightly underestimated:
- Difference (actual_range_pct - vix_predicted_move_pct) between 0.0% and 0.5%
- These are cases where VIX was "almost right" but missed slightly

Requires: openpyxl for Excel formatting
"""

import pandas as pd
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("Install with: pip3 install openpyxl")
    sys.exit(1)


def format_vix_results(csv_file: str, output_file: str = None):
    """
    Read CSV, apply conditional formatting, and export to Excel

    Args:
        csv_file: Input CSV file path
        output_file: Output Excel file path (default: same name with .xlsx)
    """
    # Default output filename
    if output_file is None:
        output_file = csv_file.replace('.csv', '_formatted.xlsx')

    print(f"📖 Reading CSV: {csv_file}")

    # Read CSV
    df = pd.read_csv(csv_file)

    # Calculate difference if not already present
    if 'diff_pct' not in df.columns:
        if 'actual_range_pct' in df.columns and 'vix_predicted_move_pct' in df.columns:
            df['diff_pct'] = df['actual_range_pct'] - df['vix_predicted_move_pct']
            print("ℹ️  Calculated diff_pct column")
        else:
            print("❌ Error: Required columns not found (actual_range_pct, vix_predicted_move_pct)")
            sys.exit(1)

    # Export to Excel first
    print(f"📝 Exporting to Excel: {output_file}")
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Load workbook for formatting
    print("🎨 Applying conditional formatting...")
    wb = load_workbook(output_file)
    ws = wb.active

    # Define green fill
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Find column indices
    headers = [cell.value for cell in ws[1]]
    diff_col_idx = headers.index('diff_pct') + 1 if 'diff_pct' in headers else None

    if diff_col_idx is None:
        print("❌ Error: diff_pct column not found in Excel")
        wb.close()
        return

    # Count rows that will be highlighted
    highlight_count = 0

    # Apply formatting (start from row 2, skip header)
    for row_idx in range(2, ws.max_row + 1):
        diff_value = ws.cell(row=row_idx, column=diff_col_idx).value

        # Check if difference is between 0.0 and 0.5
        if diff_value is not None and 0.0 < diff_value <= 0.5:
            # Apply green background to entire row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = green_fill
            highlight_count += 1

    # Save workbook
    wb.save(output_file)
    wb.close()

    print(f"✅ Formatting complete!")
    print(f"   Total rows: {len(df)}")
    print(f"   Highlighted rows (0.0 < diff ≤ 0.5): {highlight_count}")
    print(f"   Output: {output_file}")

    # Show summary of difference distribution
    print("\n📊 Difference Distribution:")
    print(f"   Negative (VIX overestimated): {(df['diff_pct'] < 0).sum()}")
    print(f"   0.0 - 0.2%: {((df['diff_pct'] > 0) & (df['diff_pct'] <= 0.2)).sum()}")
    print(f"   0.2 - 0.5%: {((df['diff_pct'] > 0.2) & (df['diff_pct'] <= 0.5)).sum()}")
    print(f"   > 0.5%: {(df['diff_pct'] > 0.5).sum()}")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Format VIX backtest CSV with conditional coloring',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Format V5 results
  python3 format_excel_with_colors.py vix_all_expiries_results_v5.csv

  # Format V4 results with custom output name
  python3 format_excel_with_colors.py vix_all_expiries_results_v4.csv -o v4_highlighted.xlsx

  # Format V2 results
  python3 format_excel_with_colors.py vix_all_expiries_results_v2.csv

Green highlight applied when:
  - Difference (actual - predicted) is between 0.0% and 0.5%
  - These are cases where VIX was "almost right" but slightly underestimated
        """
    )

    parser.add_argument('csv_file', help='Input CSV file path')
    parser.add_argument('-o', '--output', help='Output Excel file path (default: <input>_formatted.xlsx)')

    args = parser.parse_args()

    # Check if input file exists
    if not Path(args.csv_file).exists():
        print(f"❌ Error: File not found: {args.csv_file}")
        sys.exit(1)

    # Format the file
    format_vix_results(args.csv_file, args.output)


if __name__ == "__main__":
    main()
